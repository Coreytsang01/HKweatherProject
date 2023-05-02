## %% import module and define variable  
import openpyxl
import os
import pandas as pd
worksheetname = ["Chek Lap Kok","Cheung Chau","Clear Water Bay","Happy Valley","HK Observatory","HK Park","Kai Tak Runway Park","Kau Sai Chau","King's Park","Kowloon City","Kwun Tong","Lau Fau Shan","Ngong Ping","Pak Tam Chung","Peng Chau","Sai Kung","Sha Tin","Sham Shui Po","Shau Kei Wan","Shek Kong","Sheung Shui","Stanley","Ta Kwu Ling","Tai Lung","Tai Mei Tuk","Tai Mo Shan","Tai Po","Tate's Cairn","The Peak","Tseung Kwan O","Tsing Yi","Tsuen Wan Ho Koon","Tsuen Wan Shing Mun Valley","Tuen Mun","Waglan Island","Wetland Park","Wong Chuk Hang","Wong Tai Sin","Yuen Long Park",
]
check_file_exist = 'Master Data.xlsx'
column_df = pd.DataFrame(columns=[
                                        'Date', 
                                        'Temperature (deg. C)', 
                                        'Relative Humidity (%)', 
                                        'Pressure (hPa)', 
                                        'Wind Direction (Deg.)', 
                                        'Wind Speed (km/h)',
                                        'Maxiumn Gust (km/h)'
                                    ])
## %% Generate excel file 'Master Data.xlsx' if it does not existed ===============
if os.path.exists(check_file_exist) == False:
        column_df.to_excel('Master Data.xlsx')
        print('the file is generated')
else:
        print('the file already exists')

## %% Generate multiple worksheet with specific name ==============================
workbook = openpyxl.load_workbook("Master Data.xlsx")
for worksheet in worksheetname:
        if worksheet not in workbook.sheetnames:
                workbook.create_sheet(worksheet)
        else:
                print('the worksheet %s is already existed' %worksheet)
                

if 'Sheet1' in workbook.sheetnames:
                workbook.remove(workbook['Sheet1'])

workbook.save("Master Data.xlsx")

## %% Insert the column header in every worksheet if it does not exist ============
for locations in worksheetname:
        workbook.active = workbook[locations]
        df_workbook_Master = pd.read_excel('Master Data.xlsx',sheet_name=locations)
        if list(df_workbook_Master.columns) == list(column_df) :
                print('The column header for the worksheet (%s) already exist' %locations)
        else:
                with pd.ExcelWriter(
                "Master Data.xlsx",
                mode="a",
                engine="openpyxl",
                if_sheet_exists="replace",
                ) as writer:
                        column_df.to_excel(writer, sheet_name=locations,index=False)
print('Finish all process')
