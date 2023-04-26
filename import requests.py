#%%
import requests
import os
import pandas as pd
import openpyxl
import csv

#%%
downloadUrl = [
'https://data.weather.gov.hk/weatherAPI/hko_data/regional-weather/latest_10min_wind.csv', 
'https://data.weather.gov.hk/weatherAPI/hko_data/regional-weather/latest_1min_temperature.csv',
'https://data.weather.gov.hk/weatherAPI/hko_data/regional-weather/latest_1min_pressure.csv',
'https://data.weather.gov.hk/weatherAPI/hko_data/regional-weather/latest_1min_humidity.csv'
]
downloadfile_directory = 'temp file'

for url in downloadUrl:
    response = requests.get(url)
    if response.status_code == 200:
        file_path = os.path.join(downloadfile_directory, os.path.basename(url))
        with open(file_path, 'wb') as f:
            f.write(response.content)

#%%

workbook_Master = openpyxl.load_workbook('Master Data.xlsx')
workbook_Master.sheetnames[0]
df_workbook_Master = pd.read_excel('Master Data.xlsx')
print(df_workbook_Master)

df_wind = pd.read_csv('temp file\latest_1min_pressure.csv')
print(df_wind)
wind_column_list = list(df_wind.columns)
print(wind_column_list)

# # %%
# #   Date cleaning
# wind_data_cols = [
#     'Date',
#     'Location',
#     'Wind Direction',
#     'Wind Mean Speed',
#     'Maximum Gust'
# ]
# #new_df_wind = df[wind_data_cols]

# date = df_wind['Date time']
# for i in date:
#     x = str(i)
#     x1 = [x[6:8],'/',x[4:6],'/',x[:4],' ',x[-4:-2],':',x[-2:]]
#     x1 = ''.join(x)
#     print(x1)
# %%
#   Combine all temporary file into one excel
df_wind = pd.read_csv('temp file\latest_10min_wind.csv')
df_pressure = pd.read_csv('temp file\latest_1min_pressure.csv')
df_temperature = pd.read_csv('temp file\latest_1min_temperature.csv')
df_humidity = pd.read_csv('temp file\latest_1min_humidity.csv')
concat_df1 = df_temperature.merge(df_wind, how = 'left', on = ['Date time', 'Automatic Weather Station'])
concat_df2 = concat_df1.merge(df_pressure, how = 'left', on = ['Date time', 'Automatic Weather Station'])
final_combine_df = concat_df2.merge(df_humidity, how = 'left', on = ['Date time', 'Automatic Weather Station'])
final_combine_df["Date time"] = final_combine_df["Date time"].astype(str)
date = final_combine_df['Date time']
for i in date:
    x = str(i)
    x1 = [x[6:8],'/',x[4:6],'/',x[:4],' ',x[-4:-2],':',x[-2:]]
    x2 = ''.join(x1)
    print(x2)
    final_combine_df['Date time'] = final_combine_df['Date time'].replace(to_replace=i, value=x2)

print(final_combine_df)
final_combine_df.to_excel("test.xlsx", index = False )


# %%
# Try to do: insert data into master data.xlsx

combine_excel = pd.read_excel('test.xlsx', index_col='Automatic Weather Station')
workbook_Master = openpyxl.load_workbook('Master Data.xlsx')
#print(combine_excel)

# number of station: 39
Station_location = [
"Chek Lap Kok","Cheung Chau","Clear Water Bay",
"Happy Valley","HK Observatory","HK Park",
"Kai Tak Runway Park","Kau Sai Chau","King's Park",
"Kowloon City","Kwun Tong","Lau Fau Shan",
"Ngong Ping","Pak Tam Chung","Peng Chau",
"Sai Kung","Sha Tin","Sham Shui Po",
"Shau Kei Wan","Shek Kong","Sheung Shui",
"Stanley","Ta Kwu Ling","Tai Lung",
"Tai Mei Tuk","Tai Mo Shan","Tai Po",
"Tate's Cairn","The Peak","Tseung Kwan O",
"Tsing Yi","Tsuen Wan Ho Koon","Tsuen Wan Shing Mun Valley",
"Tuen Mun","Waglan Island","Wetland Park",
"Wong Chuk Hang","Wong Tai Sin","Yuen Long Park"
]
workbook_Master = openpyxl.load_workbook('Master Data.xlsx')
# for i in range(38):
#     if 
# %%
